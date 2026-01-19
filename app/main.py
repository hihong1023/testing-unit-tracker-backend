from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Optional, Dict, Any, Tuple
from datetime import datetime, timedelta, timezone
SENTINEL_FINISHED_AT = datetime(1970, 1, 1, tzinfo=timezone.utc)
from uuid import uuid4
from pathlib import Path
import hashlib
import io
import re
from fastapi import BackgroundTasks
import os
import json
import urllib.request
import urllib.error
import urllib.parse

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font

from sqlmodel import Session, select

# 🔹 IMPORTANT: import everything you need from db.py here
from app.db import (
    engine,
    get_session,
    download_db_from_blob_if_needed,
    upload_db_to_blob,
    init_db,
)
# 🔹 Your models
from app.models import (
    Unit,
    Assignment,
    Result,
    FileMeta,
    Notification,
    Token,
    User,
    TestStep,
    AssignmentUpdate,
)

# =====================================================
# App & CORS
# =====================================================

app = FastAPI(title="Testing Unit Tracker")

origins = [
    "https://calm-field-0fa08ef00.1.azurestaticapps.net",
    "http://localhost:5173",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,        # you can later tighten to just your frontend origins
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition"],
)


@app.on_event("startup")
def on_startup():
    # 1) Ensure we have the latest DB from blob
    download_db_from_blob_if_needed()

    # 2) Ensure tables exist
    init_db()

    print("[APP] Startup complete.")


@app.on_event("shutdown")
def on_shutdown():
    # Upload the current DB to blob so it persists
    upload_db_to_blob()
    print("[APP] Shutdown complete.")

# =====================================================
# Auth (simple token-based)
# =====================================================

class LoginRequest(BaseModel):
    name: str  # username from frontend

class LoginResponse(BaseModel):
    access_token: str
    role: str
    user: User

security = HTTPBearer()
TOKEN_TTL = timedelta(hours=12)

async def get_current_user(
    creds: HTTPAuthorizationCredentials = Depends(security),
    session: Session = Depends(get_session),
) -> User:
    token_str = creds.credentials
    token_row = session.get(Token, token_str)

    if not token_row:
        raise HTTPException(status_code=401, detail="Invalid or expired token")

    if datetime.utcnow() - token_row.issued_at > TOKEN_TTL:
        session.delete(token_row)
        session.commit()
        raise HTTPException(status_code=401, detail="Token expired")

    return User(id=token_row.user_id, name=token_row.name, role=token_row.role)

def require_role(required_role: str):
    async def dep(user: User = Depends(get_current_user)) -> User:
        if user.role != required_role:
            # hide existence of the route from wrong-role callers
            raise HTTPException(status_code=404, detail="Not found")
        return user
    return dep

# Preset accounts (flat list for login, etc.)
PRESET_TESTERS = [
    "alex",
    "brian",
    "ge fan",
    "jimmy",
    "kae",
    "krishnan",   # use whichever spelling you actually want
    "nicholas",
    "sunny",
    "yew meng",
    "yubo",       # <-- this one will be "individual only"
    "zhen yang",
    "sook huy",
]

PRESET_SUPERVISORS = ["kian siang", "alban", "hai hong"]

# Optional: logical groups of testers
# (you can adjust this mapping any time)
PRESET_TESTER_GROUPS: Dict[str, List[str]] = {
    "Physical Layer": [
        "sunny",
        "zhen yang",
        
    ],
    "Reliability": [
        "brian",
        "nicholas",
        "yew meng",
    ],

    "OAT":[
        "krishnan",
        "sook huy",
    ],
}


def groups_for_tester(name: str) -> List[str]:
    """
    Return list of group names that this tester belongs to.
    If tester is not in any group (e.g. 'yubo'), returns [].
    """
    name_l = name.lower()
    groups: List[str] = []
    for g, members in PRESET_TESTER_GROUPS.items():
        if any(m.lower() == name_l for m in members):
            groups.append(g)
    return groups


def group_id(group_name: str) -> str:
    """
    Uniform way to encode 'group assignment' in the tester_id field.
    Example: 'RF' -> 'group:RF'
    """
    return f"group:{group_name}"


@app.post("/auth/login", response_model=LoginResponse)
def login(body: LoginRequest, session: Session = Depends(get_session)):
    username_raw = body.name.strip()
    if not username_raw:
        raise HTTPException(status_code=400, detail="Name is required")

    key = username_raw.lower()

    if key in PRESET_SUPERVISORS:
        role = "supervisor"
    elif key in PRESET_TESTERS:
        role = "tester"
    else:
        raise HTTPException(status_code=401, detail="Unknown user")

    # 🔹 Always store normalized lowercase name
    user = User(id=str(uuid4()), name=key, role=role)

    token_str = str(uuid4())
    token_row = Token(
        token=token_str,
        user_id=user.id,
        name=user.name,
        role=user.role,
    )
    session.add(token_row)
    session.commit()

    return LoginResponse(access_token=token_str, role=role, user=user)


@app.get("/testers", response_model=List[str])
def list_testers(supervisor: User = Depends(require_role("supervisor"))):
    """Scheduler uses this to populate tester dropdown (supervisor only)."""
    return PRESET_TESTERS

@app.get("/testers/groups")
def list_tester_groups(supervisor: User = Depends(require_role("supervisor"))):
    """
    Return mapping of group_name -> list of tester names.
    Frontend can use this to build 'group' options like 'RF (group)'.
    """
    return PRESET_TESTER_GROUPS


# =====================================================
# Teams Notification
# =====================================================
FRONTEND_URL = os.getenv("FRONTEND_URL", "").rstrip("/")


def _slug_env(s: str) -> str:
    s = (s or "").strip().upper()
    s = re.sub(r"[^A-Z0-9]+", "_", s)
    return s.strip("_")


def _teams_webhooks_for_assignee(tester_id: str | None) -> list[str]:
    """
    Returns webhook URLs for a given assignee.
    Env vars supported:

    - TEAMS_WEBHOOK_DEFAULT
    - TEAMS_WEBHOOK_TESTER_<NAME>     e.g. TEAMS_WEBHOOK_TESTER_ALEX
    - TEAMS_WEBHOOK_GROUP_<GROUP>     e.g. TEAMS_WEBHOOK_GROUP_PHYSICAL_LAYER

    Assignee formats:
    - "alex"
    - "group:Physical Layer"
    """
    if not tester_id:
        return []

    urls: list[str] = []

    if tester_id.startswith("group:"):
        group_name = tester_id.split(":", 1)[1].strip()
        url = os.getenv(f"TEAMS_WEBHOOK_GROUP_{_slug_env(group_name)}")
        if url:
            urls.append(url)
    else:
        # Optional per-tester routing (if you create these env vars)
        per_user = os.getenv(f"TEAMS_WEBHOOK_TESTER_{_slug_env(tester_id)}")
        if per_user:
            urls.append(per_user)

    # Always fallback to default if nothing matched
    if not urls:
        default_url = os.getenv("TEAMS_WEBHOOK_DEFAULT")
        if default_url:
            urls.append(default_url)

    # de-dupe
    return list(dict.fromkeys([u for u in urls if u]))


def send_teams_message(webhook_url: str, title: str, lines: list[str]) -> None:
    """
    Sends an Office365 Connector MessageCard to Teams Incoming Webhook / Workflows.
    """
    payload = {
        "@type": "MessageCard",
        "@context": "http://schema.org/extensions",
        "summary": title,
        "themeColor": "0078D7",
        "title": title,
        "text": "<br/>".join(lines),
    }

    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        webhook_url,
        data=data,
        headers={"Content-Type": "application/json"},
        method="POST",
    )

    with urllib.request.urlopen(req, timeout=6) as resp:
        _ = resp.read()



TEAMS_FLOW_DM_URL = os.getenv("TEAMS_FLOW_DM_URL", "").strip()

def _slug_env(s: str) -> str:
    s = (s or "").strip().upper()
    s = re.sub(r"[^A-Z0-9]+", "_", s)
    return s.strip("_")

def tester_upn(name: str) -> str | None:
    # name like "alex" or "zhen yang"
    key = f"TESTER_UPN_{_slug_env(name)}"
    return os.getenv(key)

def assignee_upns(tester_id: str | None) -> list[tuple[str, str]]:
    """
    Returns list of (display_name, upn) recipients.
    Supports:
      - "alex"
      - "group:Physical Layer"  -> expands to all members in that group
    """
    if not tester_id:
        return []

    recips: list[tuple[str, str]] = []

    if tester_id.startswith("group:"):
        group_name = tester_id.split(":", 1)[1].strip()
        members = PRESET_TESTER_GROUPS.get(group_name, [])
        for m in members:
            upn = tester_upn(m)
            if upn:
                recips.append((m, upn))
    else:
        upn = tester_upn(tester_id)
        if upn:
            recips.append((tester_id, upn))

    # de-dupe
    seen = set()
    out: list[tuple[str, str]] = []
    for name, upn in recips:
        k = upn.lower()
        if k not in seen:
            seen.add(k)
            out.append((name, upn))
    return out

def send_flow_dm(payload: dict) -> None:
    """
    Calls your Power Automate 'When an HTTP request is received' URL.
    """
    if not TEAMS_FLOW_DM_URL:
        print("[TEAMS] TEAMS_FLOW_DM_URL not set, skipping DM.")
        return

    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        TEAMS_FLOW_DM_URL,
        data=data,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=10) as resp:
        _ = resp.read()

# =====================================================
# Static Test Steps
# =====================================================

STEPS: List[TestStep] = [
    TestStep(id=1, name="Box Build", order=1),
    TestStep(id=2, name="Connectivity Test", order=2),
    TestStep(id=3, name="Functionality Test", order=3),
    TestStep(id=4, name="EIRP Determination & Stability Calibration", order=4),
    TestStep(id=5, name="Pre-Vibration Physical Layer Test", order=5),
    TestStep(id=6, name="Vibration Test", order=6),
    TestStep(id=7, name="Post-Vibration Physical Layer Test", order=7),
    TestStep(id=8, name="Thermal Cycling", order=8),
    TestStep(id=9, name="Post-Thermal Cycling Physical Layer Test", order=9),
    TestStep(id=10, name="Burn-in Test", order=10),
    TestStep(id=11, name="EMI/EMC Test", order=11),
    TestStep(id=12, name="Post-EMI/EMC Physical Layer Test", order=12),
    TestStep(id=13, name="BGAN Network Emulator Test", order=13),
    TestStep(id=14, name="Over-the-Air Test", order=14),
    TestStep(id=15, name="OQC", order=15),
    TestStep(id=16, name="Packaging", order=16),
]

STEP_BY_ID: Dict[int, TestStep] = {s.id: s for s in STEPS}
STEP_IDS_ORDERED = [s.id for s in sorted(STEPS, key=lambda s: s.order)]

@app.get("/steps", response_model=List[TestStep])
def list_steps(user: User = Depends(get_current_user)):
    return STEPS

# =====================================================
# Units
# =====================================================

class CreateUnitRequest(BaseModel):
    unit_id: str
    sku: Optional[str] = None
    rev: Optional[str] = None
    lot: Optional[str] = None

class UnitSummary(BaseModel):
    unit_id: str
    status: str
    progress_percent: float
    passed_steps: int
    total_steps: int
    next_step_id: Optional[int]
    next_step_name: Optional[str]

class UnitDetails(BaseModel):
    unit: Unit
    assignments: List[Assignment]
    results: List[Result]

@app.post("/units", response_model=Unit)
def create_unit(
    body: CreateUnitRequest,
    supervisor: User = Depends(require_role("supervisor")),
    session: Session = Depends(get_session),
):
    unit_id = body.unit_id.strip()
    if not unit_id:
        raise HTTPException(status_code=400, detail="unit_id required")

    existing = session.get(Unit, unit_id)
    if existing:
        raise HTTPException(status_code=400, detail="Unit already exists")

    unit = Unit(
        id=unit_id,
        sku=body.sku,
        rev=body.rev,
        lot=body.lot,
        current_step_id=STEP_IDS_ORDERED[0] if STEP_IDS_ORDERED else None,
    )
    session.add(unit)

    # create assignments for all steps
    for idx, step_id in enumerate(STEP_IDS_ORDERED):
        a = Assignment(
            unit_id=unit_id,
            step_id=step_id,
            prev_passed=(idx == 0),  # first step unblocked
        )
        session.add(a)

    session.commit()
    session.refresh(unit)
    return unit

@app.delete("/units/{unit_id}", status_code=200)
def delete_unit(
    unit_id: str,
    supervisor: User = Depends(require_role("supervisor")),
    session: Session = Depends(get_session),
):
    """
    Delete a unit and all related data.
    Idempotent: if unit doesn't exist, returns 200 with deleted=False.
    """
    unit = session.get(Unit, unit_id)
    if not unit:
        return {"ok": True, "deleted": False}

    # delete file metadata + physical files
    files = session.exec(
        select(FileMeta).where(FileMeta.unit_id == unit_id)
    ).all()
    for f in files:
        if f.stored_path:
            p = Path(f.stored_path)
            try:
                p.unlink()
            except FileNotFoundError:
                pass
        session.delete(f)

    # delete results
    results = session.exec(
        select(Result).where(Result.unit_id == unit_id)
    ).all()
    for r in results:
        session.delete(r)

    # delete assignments
    assignments = session.exec(
        select(Assignment).where(Assignment.unit_id == unit_id)
    ).all()
    for a in assignments:
        session.delete(a)

    # delete notifications for this unit
    notes = session.exec(
        select(Notification).where(Notification.unit_id == unit_id)
    ).all()
    for n in notes:
        session.delete(n)

    session.delete(unit)
    session.commit()

    return {"ok": True, "deleted": True}

class RenameUnitRequest(BaseModel):
    new_unit_id: str


@app.patch("/units/{unit_id}/rename", response_model=Unit)
def rename_unit(
    unit_id: str,
    body: RenameUnitRequest,
    supervisor: User = Depends(require_role("supervisor")),
    session: Session = Depends(get_session),
):
    """
    Rename a unit and update all related records (assignments, results,
    files, notifications) to use the new unit_id.

    Frontend:
      PATCH /units/{unit_id}/rename
      { "new_unit_id": "NEW_ID" }
    """
    new_id = body.new_unit_id.strip()
    if not new_id:
        raise HTTPException(status_code=400, detail="new_unit_id required")

    if new_id == unit_id:
        # nothing to do
        unit = session.get(Unit, unit_id)
        if not unit:
            raise HTTPException(status_code=404, detail="Unit not found")
        return unit

    # Check if target id already exists
    existing = session.get(Unit, new_id)
    if existing:
        raise HTTPException(status_code=400, detail="Target unit_id already exists")

    # Load the unit we are renaming
    unit = session.get(Unit, unit_id)
    if not unit:
        raise HTTPException(status_code=404, detail="Unit not found")

    # --- Update related tables first ---

    # Assignments
    assignments = session.exec(
        select(Assignment).where(Assignment.unit_id == unit_id)
    ).all()
    for a in assignments:
        a.unit_id = new_id
        session.add(a)

    # Results
    results = session.exec(
        select(Result).where(Result.unit_id == unit_id)
    ).all()
    for r in results:
        r.unit_id = new_id
        session.add(r)

    # File metadata
    files = session.exec(
        select(FileMeta).where(FileMeta.unit_id == unit_id)
    ).all()
    for f in files:
        f.unit_id = new_id
        session.add(f)
        # (Optional) we are NOT renaming physical file paths here;
        # stored_path can still contain the old unit id in the filename.

    # Notifications (if any)
    notes = session.exec(
        select(Notification).where(Notification.unit_id == unit_id)
    ).all()
    for n in notes:
        n.unit_id = new_id
        session.add(n)

    # --- Finally update the Unit primary key ---
    unit.id = new_id
    session.add(unit)

    session.commit()
    session.refresh(unit)
    return unit

@app.get("/units/summary", response_model=List[UnitSummary])
def get_units_summary(
    user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    units = session.exec(select(Unit)).all()
    summaries: List[UnitSummary] = []

    for u in units:
        # All assignments for this unit, sorted by step order
        assigns = session.exec(
            select(Assignment).where(Assignment.unit_id == u.id)
        ).all()
        assigns.sort(key=lambda a: STEP_BY_ID[a.step_id].order)

        # 🔹 Only count non-skipped steps in progress
        effective_assigns = [a for a in assigns if not a.skipped]
        total_steps = len(effective_assigns)

        # All results for this unit
        results = session.exec(
            select(Result).where(Result.unit_id == u.id)
        ).all()
        result_map = {(r.unit_id, r.step_id): r for r in results}

        # 🔹 Passed steps = number of effective steps with a PASS result
        passed_steps = sum(
            1
            for a in effective_assigns
            if (u.id, a.step_id) in result_map
            and result_map[(u.id, a.step_id)].passed
        )

        # 🔹 Progress percentage (0–100)
        if total_steps > 0:
            progress = (passed_steps / total_steps) * 100.0
        else:
            progress = 0.0

        # 🔹 Next step: first effective step that does NOT have a result
        next_step_id: Optional[int] = None
        next_step_name: Optional[str] = None
        for a in effective_assigns:
            if (u.id, a.step_id) not in result_map:
                next_step_id = a.step_id
                next_step_name = STEP_BY_ID[a.step_id].name
                break

        # 🔹 Status
        if total_steps == 0:
            status = "EMPTY"
        else:
            # How many effective steps actually have a result?
            completed_effective = sum(
                1 for a in effective_assigns if (u.id, a.step_id) in result_map
            )
            if completed_effective == total_steps:
                status = "COMPLETED"
            else:
                status = "IN_PROGRESS"

        summaries.append(
            UnitSummary(
                unit_id=u.id,
                status=status,
                progress_percent=progress,
                passed_steps=passed_steps,
                total_steps=total_steps,
                next_step_id=next_step_id,
                next_step_name=next_step_name,
            )
        )

    return summaries


@app.get("/units/{unit_id}/details", response_model=UnitDetails)
def get_unit_details(
    unit_id: str,
    user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    """
    Details view for a single unit.
    Used by UnitsPage details route: /units/{unit_id}/details
    """
    unit = session.get(Unit, unit_id)
    if not unit:
        raise HTTPException(status_code=404, detail="Unit not found")

    assignments = session.exec(
        select(Assignment).where(Assignment.unit_id == unit_id)
    ).all()
    assignments.sort(key=lambda a: STEP_BY_ID[a.step_id].order)

    results = session.exec(
        select(Result).where(Result.unit_id == unit_id)
    ).all()

    return UnitDetails(unit=unit, assignments=assignments, results=results)

# =====================================================
# Results & Uploads
# =====================================================
def recompute_unit_status(session: Session, unit_id: str) -> None:
    """
    Helper: set unit.status to COMPLETED when all assignments are
    either skipped or finished (DONE/PASS/FAIL). Otherwise mark as IN_PROGRESS.
    """
    unit = session.get(Unit, unit_id)
    if not unit:
        return

    assignments = session.exec(
        select(Assignment).where(Assignment.unit_id == unit_id)
    ).all()

    if not assignments:
        # no assignments, keep whatever status it has
        return

    all_done = all(
        a.skipped or a.status in ("DONE", "PASS", "FAIL")
        for a in assignments
    )

    if all_done:
        unit.status = "COMPLETED"
    else:
        # only override if previously completed; otherwise keep existing
        if unit.status == "COMPLETED":
            unit.status = "IN_PROGRESS"

    session.add(unit)
    
class ResultIn(BaseModel):
    unit_id: str
    step_id: int
    metrics: Optional[Dict[str, Any]] = None
    passed: bool
    finished_at: Optional[datetime] = None

class ResultOut(BaseModel):
    id: str
    unit_id: str
    step_id: int
    passed: bool
    metrics: Dict[str, Any]
    files: List[str]
    submitted_by: Optional[str]
    finished_at: Optional[datetime] = None   # ✅ was datetime


@app.post("/results", response_model=ResultOut)
def create_or_update_result(
    body: ResultIn,
    background_tasks: BackgroundTasks,
    user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    if body.step_id not in STEP_BY_ID:
        raise HTTPException(status_code=400, detail="Unknown step_id")

    unit = session.get(Unit, body.unit_id)
    if not unit:
        raise HTTPException(status_code=404, detail="Unit not found")

    unit_id = body.unit_id
    step_id = body.step_id
    passed = body.passed
    metrics = body.metrics or {}
    now = datetime.now(timezone.utc)

    # ✅ finished_at rules:
    # - tester queue (tester role) with no finished_at => use NOW
    # - upload page (supervisor) with no finished_at => keep None (so UI shows "-")
    if body.finished_at is None:
        finished = now if user.role == "tester" else None
    else:
        finished = body.finished_at



    # ----- create / update Result -----
    existing_result = session.exec(
        select(Result).where(
            Result.unit_id == unit_id,
            Result.step_id == step_id,
        )
    ).first()

    if existing_result:
        existing_result.metrics = metrics
        existing_result.passed = passed
    
        # ✅ Only overwrite if frontend actually provided a finished_at
        if body.finished_at is not None:
            existing_result.finished_at = body.finished_at
        else:
            # ✅ tester quick result should stamp NOW if existing was None or sentinel
            if user.role == "tester":
                if (existing_result.finished_at is None) or (existing_result.finished_at.date() == SENTINEL_FINISHED_AT.date()):
                    existing_result.finished_at = now

    
        res = existing_result
    else:
        res = Result(
            unit_id=unit_id,
            step_id=step_id,
            passed=passed,
            metrics=metrics,
            files=[],
            submitted_by=user.id,
            finished_at=finished, 
        )
        session.add(res)


    # ----- update Assignment for this step -----
    a = session.exec(
        select(Assignment).where(
            Assignment.unit_id == unit_id,
            Assignment.step_id == step_id,
        )
    ).first()
    if a:
        a.tester_id = user.name
        # store PASS/FAIL so scheduler + matrix see same state
        a.status = "PASS" if passed else "FAIL"
        session.add(a)

    # ----- chain prev_passed to next step -----
    if step_id in STEP_IDS_ORDERED:
        idx = STEP_IDS_ORDERED.index(step_id)
        if idx + 1 < len(STEP_IDS_ORDERED):
            next_step_id = STEP_IDS_ORDERED[idx + 1]
            nxt = session.exec(
                select(Assignment).where(
                    Assignment.unit_id == unit_id,
                    Assignment.step_id == next_step_id,
                )
            ).first()
            if nxt:
                nxt.prev_passed = passed
                session.add(nxt)

    # ----- recompute unit.status (COMPLETED / IN_PROGRESS) -----
    recompute_unit_status(session, unit_id)

    session.commit()
    session.refresh(res)

    # ----- create notification for next tester (DB + Teams) -----
    if passed and step_id in STEP_IDS_ORDERED:
        idx = STEP_IDS_ORDERED.index(step_id)
        if idx + 1 < len(STEP_IDS_ORDERED):
            next_step_id = STEP_IDS_ORDERED[idx + 1]
    
            next_assign = session.exec(
                select(Assignment).where(
                    Assignment.unit_id == unit_id,
                    Assignment.step_id == next_step_id,
                )
            ).first()
    
            if next_assign and next_assign.tester_id:
                dup = session.exec(
                    select(Notification).where(
                        Notification.tester_id == next_assign.tester_id,
                        Notification.unit_id == unit_id,
                        Notification.from_step_id == step_id,
                        Notification.to_step_id == next_step_id,
                    )
                ).first()
    
                if not dup:
                    msg = (
                        f"Unit {unit_id} is ready for "
                        f"{STEP_BY_ID[next_step_id].name} (previous step passed)."
                    )
    
                    note = Notification(
                        tester_id=next_assign.tester_id,
                        unit_id=unit_id,
                        from_step_id=step_id,
                        to_step_id=next_step_id,
                        message=msg,
                    )
                    session.add(note)
                    session.commit()
    
                    frontend_url = os.getenv("FRONTEND_URL", "").rstrip("/")
                    unit_link = f"{frontend_url}/units/{urllib.parse.quote(unit_id)}" if frontend_url else ""
    
                    # Build recipients (individual or expand group)
                    recips = assignee_upns(next_assign.tester_id)
    
                    for display_name, upn in recips:
                        payload = {
                            "unit_id": unit_id,
                            "from_step": STEP_BY_ID[step_id].name,
                            "to_step": STEP_BY_ID[next_step_id].name,
                            "assignee": display_name,
                            "assignee_upn": upn,
                            "unit_url": unit_link,
                            "message": msg,
                        }
    
                        # ✅ send async so your API doesn't slow down / fail
                        background_tasks.add_task(send_flow_dm, payload)


    return res



# Storage config for uploads
STORAGE_ROOT = Path("storage")
STORAGE_ROOT.mkdir(exist_ok=True)

ALLOWED_EXT = {".zip", ".csv", ".pdf", ".png"}
MAX_SIZE_BYTES = 100 * 1024 * 1024  # 100 MB

@app.post("/uploads")
async def upload_evidence(
    unit_id: str = Form(...),
    step_id: int = Form(...),
    result_id: str = Form(...),
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    unit = session.get(Unit, unit_id)
    if not unit:
        raise HTTPException(status_code=404, detail="Unit not found")

    if step_id not in STEP_BY_ID:
        raise HTTPException(status_code=400, detail="Unknown step_id")

    res = session.get(Result, result_id)
    if not res:
        raise HTTPException(status_code=404, detail="Result not found")

    if res.unit_id != unit_id or res.step_id != step_id:
        raise HTTPException(
            status_code=400,
            detail="result_id does not match unit/step",
        )

    ext = Path(file.filename).suffix.lower()
    if ext not in ALLOWED_EXT:
        raise HTTPException(status_code=400, detail=f"File type {ext} not allowed")

    content = await file.read()
    size = len(content)
    if size > MAX_SIZE_BYTES:
        raise HTTPException(status_code=400, detail="File too large")

    sha = hashlib.sha256(content).hexdigest()

    existing = session.exec(
        select(FileMeta).where(
            FileMeta.sha256 == sha,
            FileMeta.unit_id == unit_id,
            FileMeta.step_id == step_id,
        )
    ).first()

    if existing:
        if existing.id not in res.files:
            res.files.append(existing.id)
            session.add(res)
            session.commit()
        return {"file_id": existing.id, "deduplicated": True}

    bucket = sha[:2]
    bucket_dir = STORAGE_ROOT / bucket
    bucket_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
    unique_suffix = uuid4().hex[:6]
    server_name = f"{unit_id}_{step_id}_{timestamp}_{unique_suffix}{ext}"
    stored_path = bucket_dir / server_name

    with open(stored_path, "wb") as f:
        f.write(content)

    meta = FileMeta(
        unit_id=unit_id,
        step_id=step_id,
        result_id=result_id,
        orig_name=file.filename,
        stored_name=server_name,
        stored_path=str(stored_path),
        sha256=sha,
        size=size,
    )
    session.add(meta)
    session.commit()
    session.refresh(meta)

    res.files.append(meta.id)
    session.add(res)
    session.commit()

    return {"file_id": meta.id, "deduplicated": False}

@app.get("/reports/unit/{unit_id}/traveller.xlsx")
def export_traveller_log(
    unit_id: str,
    supervisor: User = Depends(require_role("supervisor")),
    session: Session = Depends(get_session),
):
    """
    Export a traveller log for one unit as Excel.

    Layout:
    Row 1:  step headers  (2. Functionality Test, 3. EIRP..., etc.)
    Row 2:  tester names  (Tester1, Tester2, ...)
    Row 3:  finish dates  (20-Nov, 21-Nov, ...)
    Row 4:  results       (Pass / Fail, coloured)
    """
    unit = session.get(Unit, unit_id)
    if not unit:
        raise HTTPException(status_code=404, detail="Unit not found")

    # All assignments & results for this unit
    assignments = session.exec(
        select(Assignment).where(Assignment.unit_id == unit_id)
    ).all()
    results = session.exec(
        select(Result).where(Result.unit_id == unit_id)
    ).all()

    # Maps by step_id for quick lookup
    assignments_by_step = {a.step_id: a for a in assignments}
    results_by_step = {r.step_id: r for r in results}

    # Use your static STEP order
    ordered_steps = [STEP_BY_ID[sid] for sid in STEP_IDS_ORDERED]

    # --- Build Excel workbook ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Traveller"

    # Helper for formatting dates like "20-Nov"
    def fmt_date(dt):
        if not dt:
            return ""
        return dt.strftime("%d-%b")

    # Row 1: step headers
    ws.cell(row=1, column=1, value="Unit")
    for col_idx, step in enumerate(ordered_steps, start=2):
        ws.cell(
            row=1,
            column=col_idx,
            value=f"{step.order}. {step.name}",
        )

    # Row 2: tester names
    ws.cell(row=2, column=1, value="Tester")
    for col_idx, step in enumerate(ordered_steps, start=2):
        a = assignments_by_step.get(step.id)
        ws.cell(row=2, column=col_idx, value=a.tester_id if a else "")

    # Row 3: finish dates
    ws.cell(row=3, column=1, value="Date")
    for col_idx, step in enumerate(ordered_steps, start=2):
        r = results_by_step.get(step.id)
        ws.cell(row=3, column=col_idx, value=fmt_date(r.finished_at) if r else "")

    # Row 4: Pass / Fail with colours
    ws.cell(row=4, column=1, value="Result")
    for col_idx, step in enumerate(ordered_steps, start=2):
        r = results_by_step.get(step.id)
        if not r:
            continue

        val = "Pass" if r.passed else "Fail"
        cell = ws.cell(row=4, column=col_idx, value=val)

        if r.passed:
            cell.fill = PatternFill("solid", fgColor="C6EFCE")  # light green
            cell.font = Font(color="006100")
        else:
            cell.fill = PatternFill("solid", fgColor="FFC7CE")  # light red
            cell.font = Font(color="9C0006")

    # Column widths & simple alignment
    for col_idx in range(1, len(ordered_steps) + 2):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 20

    # Top-left cell with unit id
    ws.cell(row=1, column=1, value=unit_id)

    # --- Return as downloadable file ---
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{unit_id}_traveller.xlsx"
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"'
    }

    return StreamingResponse(
        buf,
        media_type=(
            "application/"
            "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers=headers,
    )


# =====================================================
# Tester: Assignments view (tester home page)
# =====================================================

@app.get("/tester/assignments", response_model=List[Assignment])
def get_tester_assignments(
    tester_id: str,
    user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    """
    List assignments that a tester can see / work on.

    Frontend calls:
      GET /tester/assignments?tester_id=<name>

    Rules:
    - Tester can only query their own assignments (tester_id must match login name)
    - Supervisor can query assignments for any tester
    - Only steps that are PENDING/RUNNING, whose previous step passed,
      and that do NOT already have a Result are returned.
    """
    # Security: testers may only query themselves
    if user.role == "tester" and user.name != tester_id:
        raise HTTPException(status_code=403, detail="Forbidden")

    # Load all assignments and results
    assignments = session.exec(select(Assignment)).all()
    results = session.exec(select(Result)).all()
    result_set = {(r.unit_id, r.step_id) for r in results}

    visible: List[Assignment] = []

    # 🔹 IDs this tester is allowed to see:
    #   - their own name (e.g. "yubo")
    #   - any groups they belong to (e.g. "group:RF")
    tester_groups = groups_for_tester(tester_id)
    visible_ids = {tester_id} | {group_id(g) for g in tester_groups}

    for a in assignments:
        # ignore completed steps
        if a.status not in ("PENDING", "RUNNING"):
            continue

        # ignore steps that already have a result
        if (a.unit_id, a.step_id) in result_set:
            continue

        # previous step must be passed
        if not a.prev_passed:
            continue

        # tester can see:
        # - unassigned steps (tester_id is None)
        # - steps assigned directly to them ("yubo")
        # - steps assigned to one of their groups ("group:RF")
        if a.tester_id is not None and a.tester_id not in visible_ids:
            continue

        visible.append(a)

    # Sort nicely for UI: by unit, then by step order
    visible.sort(key=lambda x: (x.unit_id, STEP_BY_ID[x.step_id].order))
    return visible


@app.get("/tester/schedule", response_model=List[Assignment])
def get_tester_schedule(
    tester_id: str,
    user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    """
    Full schedule view for a tester.

    Used by 'Upcoming Tests (Tester)' page.
    Shows ALL assignments assigned:
      - directly to this tester
      - or to any group they belong to
    """

    # Security: testers may only query themselves
    if user.role == "tester" and user.name != tester_id:
        raise HTTPException(status_code=403, detail="Forbidden")

    tester_groups = groups_for_tester(tester_id)
    visible_ids = {tester_id} | {group_id(g) for g in tester_groups}

    # load all assignments, then filter by tester_id
    assignments = session.exec(select(Assignment)).all()
    assignments = [a for a in assignments if a.tester_id in visible_ids]

    # Sort nicely: by start date, then unit, then step order
    assignments.sort(
        key=lambda a: (
            a.start_at or datetime.max,
            a.unit_id,
            STEP_BY_ID[a.step_id].order,
        )
    )
    return assignments

# =====================================================
# Tester: set assignment status (RUNNING / PENDING)
# =====================================================

class TesterStatusUpdate(BaseModel):
    status: str  # "RUNNING" or "PENDING"


@app.post("/tester/assignments/{assign_id}/status", response_model=Assignment)
def set_tester_assignment_status(
    assign_id: str,
    body: TesterStatusUpdate,
    user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    """
    Allow a tester (or supervisor) to mark an assignment as RUNNING or
    back to PENDING. PASS/FAIL still go through /results.
    """

    # Only testers or supervisors can call this
    if user.role not in ("tester", "supervisor"):
        raise HTTPException(status_code=403, detail="Forbidden")

    a = session.get(Assignment, assign_id)
    if not a:
        raise HTTPException(status_code=404, detail="Assignment not found")

    # If caller is a tester, make sure this assignment belongs to them
    # either directly, or via one of their groups.
    if user.role == "tester":
        tester_groups = groups_for_tester(user.name)
        allowed_ids = {user.name} | {group_id(g) for g in tester_groups}
        if a.tester_id not in allowed_ids:
            raise HTTPException(status_code=403, detail="Not your assignment")

    new_status = body.status.upper()

    if new_status not in ("RUNNING", "PENDING"):
        raise HTTPException(
            status_code=400,
            detail="Only RUNNING or PENDING allowed here",
        )

    a.status = new_status
    now = datetime.now(timezone.utc)

    # ✅ When tester starts RUNNING, store actual start timestamp once
    if new_status == "RUNNING" and a.actual_start_at is None:
        a.actual_start_at = now
    
    # (Optional) if tester reverts to PENDING, don't delete the actual time
    # if new_status == "PENDING": pass

    # Optional: when starting RUNNING and no start_at yet, auto set start time
    # if new_status == "RUNNING" and a.start_at is None:
    #     a.start_at = datetime.utcnow()

    # keep unit status consistent
    recompute_unit_status(session, a.unit_id)

    session.add(a)
    session.commit()
    session.refresh(a)
    return a

# =====================================================
# Scheduling (Supervisor)
# =====================================================

class DuplicateRequest(BaseModel):
    source_unit_id: str
    new_unit_ids: List[str]
    day_shift: int = 0

@app.post("/schedule/duplicate")
def duplicate_schedule(
    body: DuplicateRequest,
    supervisor: User = Depends(require_role("supervisor")),
    session: Session = Depends(get_session),
):
    src_id = body.source_unit_id.strip()
    if not src_id:
        raise HTTPException(status_code=400, detail="source_unit_id required")

    src_unit = session.get(Unit, src_id)
    if not src_unit:
        raise HTTPException(status_code=404, detail="Source unit not found")

    src_assignments = session.exec(
        select(Assignment).where(Assignment.unit_id == src_id)
    ).all()
    src_assignments.sort(key=lambda a: STEP_BY_ID[a.step_id].order)

    def shift(dt: Optional[datetime]) -> Optional[datetime]:
        return (dt + timedelta(days=body.day_shift)) if dt else None

    created_units: List[str] = []

    for new_id_raw in body.new_unit_ids:
        new_id = new_id_raw.strip()
        if not new_id:
            continue

        if session.get(Unit, new_id):
            raise HTTPException(status_code=400, detail=f"Unit {new_id} already exists")

        new_unit = Unit(
            id=new_id,
            sku=src_unit.sku,
            rev=src_unit.rev,
            lot=src_unit.lot,
            status="IN_PROGRESS",
            current_step_id=STEP_IDS_ORDERED[0] if STEP_IDS_ORDERED else None,
        )
        session.add(new_unit)

        for src_a in src_assignments:
            # copy assignment including status
            new_a = Assignment(
                unit_id=new_id,
                step_id=src_a.step_id,
                tester_id=src_a.tester_id,
                start_at=shift(src_a.start_at),
                end_at=shift(src_a.end_at),
                status=src_a.status,          # ✅ duplicate status
                prev_passed=src_a.prev_passed, # ✅ duplicate gating
                skipped=src_a.skipped,         # ✅ duplicate skipped
            )
            session.add(new_a)

            # ✅ If PASS/FAIL, also create Result so Units progress is correct
            st = (src_a.status or "").upper()
            if st in ("PASS", "FAIL") and not src_a.skipped:
                passed = (st == "PASS")

                # date from scheduler if exists, otherwise None
                sched_finished = new_a.end_at or new_a.start_at or SENTINEL_FINISHED_AT

                session.add(
                    Result(
                        unit_id=new_id,
                        step_id=src_a.step_id,
                        passed=passed,
                        metrics={},
                        files=[],
                        submitted_by=None,
                        finished_at=sched_finished,  # may be None (after model fix below)
                    )
                )

        created_units.append(new_id)

    session.commit()
    return {"ok": True, "created_units": created_units}

class AssignmentPatch(BaseModel):
    tester_id: Optional[str] = None
    start_at: Optional[datetime] = None
    end_at: Optional[datetime] = None
    status: Optional[str] = None
    skipped: Optional[bool] = None

def overlaps(
    a_start: Optional[datetime],
    a_end: Optional[datetime],
    b_start: Optional[datetime],
    b_end: Optional[datetime],
) -> bool:
    if not a_start or not a_end or not b_start or not b_end:
        return False
    return (a_start <= b_end) and (b_start <= a_end)

@app.get("/assignments/schedule", response_model=List[Assignment])
def get_schedule(
    supervisor: User = Depends(require_role("supervisor")),
    session: Session = Depends(get_session),
):
    """Return all assignments for the scheduler page (Gantt-like view)."""
    assignments = session.exec(select(Assignment)).all()
    assignments.sort(key=lambda a: (a.unit_id, STEP_BY_ID[a.step_id].order))
    return assignments

@app.patch("/assignments/{assign_id}", response_model=Assignment)
def patch_assignment(
    assign_id: str,
    body: AssignmentPatch,
    supervisor: User = Depends(require_role("supervisor")),
    session: Session = Depends(get_session),
):
    a = session.get(Assignment, assign_id)
    if not a:
        raise HTTPException(status_code=404, detail="Assignment not found")

    sent_fields = body.__fields_set__

    # --- if skipping, clear tester & dates and mark status ---
    if body.skipped is True:
        a.skipped = True
        a.tester_id = None
        a.start_at = None
        a.end_at = None
        a.status = "SKIPPED"
    else:
        # ✅ explicit updates (including None to clear)
        if "tester_id" in sent_fields:
            a.tester_id = body.tester_id  # None → unassign

        if "start_at" in sent_fields:
            a.start_at = body.start_at    # None → clear

        if "end_at" in sent_fields:
            a.end_at = body.end_at        # None → clear

        prev_status = a.status
        if "status" in sent_fields and body.status is not None:
            a.status = body.status

        # --- sync PASS / FAIL into Result + next.prev_passed ---
        if "status" in sent_fields and body.status in ("PASS", "FAIL"):
            passed = body.status == "PASS"
        
            # ✅ Scheduler date source (may be None)
            sched_finished = a.end_at or a.start_at or SENTINEL_FINISHED_AT
        
            r = session.exec(
                select(Result).where(
                    Result.unit_id == a.unit_id,
                    Result.step_id == a.step_id,
                )
            ).first()
        
            if r:
                r.passed = passed
        
                # ✅ DO NOT force a date
                # ✅ Only fill finished_at if it is currently empty AND scheduler has a date
                # (Upload result should remain higher priority because it sets finished_at explicitly)
                if r.finished_at is None and sched_finished is not None:
                    r.finished_at = sched_finished
        
                session.add(r)
            else:
                # ✅ create result but finished_at stays None if no scheduler dates
                r = Result(
                    unit_id=a.unit_id,
                    step_id=a.step_id,
                    passed=passed,
                    metrics={},
                    files=[],
                    submitted_by=None,
                    finished_at=sched_finished,  # ✅ can be None
                )
                session.add(r)
        
            # propagate prev_passed to next step (keep your existing code)
            if a.step_id in STEP_IDS_ORDERED:
                idx = STEP_IDS_ORDERED.index(a.step_id)
                if idx + 1 < len(STEP_IDS_ORDERED):
                    next_step_id = STEP_IDS_ORDERED[idx + 1]
                    nxt = session.exec(
                        select(Assignment).where(
                            Assignment.unit_id == a.unit_id,
                            Assignment.step_id == next_step_id,
                        )
                    ).first()
                    if nxt:
                        nxt.prev_passed = passed
                        session.add(nxt)


        # if un-skipping
        if body.skipped is False:
            a.skipped = False
            if a.status == "SKIPPED":
                a.status = "PENDING"

    # --- validate overlaps only if NOT skipped ---
    if not a.skipped:
        new_start = a.start_at
        new_end = a.end_at
        if new_start and new_end and new_end < new_start:
            raise HTTPException(
                status_code=400,
                detail="End date cannot be before start date",
            )

        others = session.exec(
            select(Assignment).where(Assignment.unit_id == a.unit_id)
        ).all()
        for other in others:
            if other.id == a.id or other.skipped:
                continue
            if overlaps(new_start, new_end, other.start_at, other.end_at):
                raise HTTPException(
                    status_code=409,
                    detail=(
                        f"Unit '{a.unit_id}' already has another test scheduled "
                        f"from {other.start_at} to {other.end_at}"
                    ),
                )

    # --- recompute unit.status after any change ---
    recompute_unit_status(session, a.unit_id)

    session.add(a)
    session.commit()
    session.refresh(a)
    return a


# =====================================================
# Traveller Log
# =====================================================
def safe_sheet_title(unit_id: str) -> str:
    # Excel sheet title must be <=31 chars and not contain [ ] : * ? / \
    title = re.sub(r'[\[\]\:\*\?\/\\]', "_", unit_id)
    if len(title) > 31:
        title = title[:31]
    return title or "Unit"


def add_traveller_sheet_for_unit(
    wb: Workbook,
    unit_id: str,
    session: Session,
    first_sheet: bool = False,
):
    unit = session.get(Unit, unit_id)
    if not unit:
        # silently skip missing units; you can also raise if you prefer
        return

    # assignments + results
    assignments = session.exec(
        select(Assignment).where(Assignment.unit_id == unit_id)
    ).all()
    results = session.exec(
        select(Result).where(Result.unit_id == unit_id)
    ).all()

    assignments_by_step = {a.step_id: a for a in assignments}
    results_by_step = {r.step_id: r for r in results}

    # use your known ordered steps
    ordered_steps: list[TestStep] = [STEP_BY_ID[sid] for sid in STEP_IDS_ORDERED]

    # create / reuse sheet
    sheet_name = safe_sheet_title(unit_id)
    if first_sheet:
        ws = wb.active
        ws.title = sheet_name
    else:
        ws = wb.create_sheet(title=sheet_name)

    def fmt_date(dt):
        if not dt or dt.date() == SENTINEL_FINISHED_AT.date():
            return "-"
        return dt.strftime("%d-%b")

    # Row 1: step headers
    ws.cell(row=1, column=1, value=unit_id)
    for col_idx, step in enumerate(ordered_steps, start=2):
        ws.cell(
            row=1,
            column=col_idx,
            value=f"{step.order}. {step.name}",
        )

    # Row 2: testers
    ws.cell(row=2, column=1, value="Tester")
    for col_idx, step in enumerate(ordered_steps, start=2):
        a = assignments_by_step.get(step.id)
        ws.cell(row=2, column=col_idx, value=a.tester_id if a else "")

    # Row 3: dates
    ws.cell(row=3, column=1, value="Date")
    for col_idx, step in enumerate(ordered_steps, start=2):
        r = results_by_step.get(step.id)
        ws.cell(row=3, column=col_idx, value=fmt_date(r.finished_at) if r else "")

    # Row 4: result
    ws.cell(row=4, column=1, value="Result")
    for col_idx, step in enumerate(ordered_steps, start=2):
        r = results_by_step.get(step.id)
        if not r:
            continue
        val = "Pass" if r.passed else "Fail"
        cell = ws.cell(row=4, column=col_idx, value=val)
        if r.passed:
            cell.fill = PatternFill("solid", fgColor="C6EFCE")
            cell.font = Font(color="006100")
        else:
            cell.fill = PatternFill("solid", fgColor="FFC7CE")
            cell.font = Font(color="9C0006")

    # column widths
    for col_idx in range(1, len(ordered_steps) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

class BulkTravellerRequest(BaseModel):
    unit_ids: List[str]


@app.post("/reports/traveller/bulk.xlsx")
def export_traveller_bulk_xlsx(
    payload: BulkTravellerRequest,
    supervisor: User = Depends(require_role("supervisor")),
    session: Session = Depends(get_session),
):
    if not payload.unit_ids:
        raise HTTPException(status_code=400, detail="No unit_ids provided")

    wb = Workbook()
    first = True
    for unit_id in payload.unit_ids:
        add_traveller_sheet_for_unit(wb, unit_id, session, first_sheet=first)
        first = False

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    headers = {
        "Content-Disposition": 'attachment; filename="traveller_logs.xlsx"'
    }

    return StreamingResponse(
        buf,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers=headers,
    )

# =====================================================
# Root
# =====================================================

@app.get("/")
def root():
    return {"message": "Testing Unit Tracker API running"}











































